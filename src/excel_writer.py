"""Output writers for opportunity ranking artifacts."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR, SCORING_COLUMNS


def ensure_output_dir(path: Path = OUTPUT_DIR) -> None:
    """Create the output directory if it does not already exist."""

    path.mkdir(parents=True, exist_ok=True)


def write_opportunity_workbook(dataframe: pd.DataFrame, output_path: Path) -> None:
    """Write the scored ranking to Excel with light formatting."""

    ensure_output_dir(output_path.parent)

    scoring_config = pd.DataFrame(
        [
            {
                "Column": column,
                "Weight": config["weight"],
                "Direction": "Lower is better" if config["inverse"] else "Higher is better",
            }
            for column, config in SCORING_COLUMNS.items()
        ]
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, sheet_name="Opportunity Ranking", index=False)
        scoring_config.to_excel(writer, sheet_name="Scoring Model", index=False)

        _format_worksheet(writer.book["Opportunity Ranking"])
        _format_worksheet(writer.book["Scoring Model"])


def write_summary_report(dataframe: pd.DataFrame, output_path: Path) -> None:
    """Generate a concise Markdown report for the scored niche database."""

    ensure_output_dir(output_path.parent)

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    scored_rows = dataframe["Opportunity Score"].notna().sum()
    average_score = dataframe["Opportunity Score"].mean()
    top_rows = dataframe.head(10)

    lines = [
        "# Adobe Stock Opportunity Summary",
        "",
        f"Generated: {generated_at}",
        "",
        "## Scoring Model",
        "",
        "Opportunity Score is calculated from available 1-10 scoring columns. Demand, Commercial Value, Evergreen, and Repeat Use are positive signals. Competition and AI Saturation are inverted so lower saturation improves the score.",
        "",
        f"- Niches analyzed: {len(dataframe)}",
        f"- Niches with scores: {scored_rows}",
        f"- Average opportunity score: {average_score:.1f}" if pd.notna(average_score) else "- Average opportunity score: n/a",
        "",
        "## Top Opportunities",
        "",
        "| Rank | Industry | Niche | Score | Priority | Buyer Persona |",
        "| ---: | --- | --- | ---: | --- | --- |",
    ]

    for index, row in top_rows.iterrows():
        lines.append(
            "| {rank} | {industry} | {niche} | {score} | {priority} | {buyer} |".format(
                rank=index + 1,
                industry=_markdown_cell(row.get("Industry")),
                niche=_markdown_cell(row.get("Niche")),
                score=_format_score(row.get("Opportunity Score")),
                priority=_markdown_cell(row.get("Calculated Priority")),
                buyer=_markdown_cell(row.get("Buyer Persona")),
            )
        )

    lines.extend(
        [
            "",
            "## Priority Distribution",
            "",
        ]
    )

    priority_counts = dataframe["Calculated Priority"].value_counts(dropna=False)
    for priority, count in priority_counts.items():
        lines.append(f"- {priority}: {count}")

    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _format_worksheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def _markdown_cell(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).replace("|", "\\|").strip()


def _format_score(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(int(value))
